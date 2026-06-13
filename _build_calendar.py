from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import date

rows = [
    # Date, Weekday, Week, Content type, Platform, Topic / Angle, Hook / Copy hint, Status, Owner
    (date(2026,5,5),  "Tue", "Wk 1 (5–10 May)", "Campaign opener",          "LinkedIn", "Why we're opening an AI Learning Cluster (and why it's not another prompting course)", "Lead with what L&D practitioners need to PROTECT — voice, judgment, credibility — before what AI unlocks. Tease the Override Log as the cluster's hero practice.", "Draft", "Sara"),
    (date(2026,5,6),  "Wed", "Wk 1 (5–10 May)", "Community heads-up",       "Slack",    "Heads up Shakers: hold Tue 2 June, 17:00 CEST",                                          "Friendly drop in #general / #ai-cluster: 'Something is brewing. Block your calendar for the Opening Expo on 2 June. Cohort kicks off the week after. More this week.'",                                            "Draft", "Sara"),

    (date(2026,5,11), "Mon", "Wk 2 (11–17 May)", "Expo signup launch",       "LinkedIn", "Save your seat — AI Learning Cluster Opening Expo, 2 June, 17:00 CEST",              "Public signup post. One image, one paragraph, one link. State who it's for and who it isn't. End with: 'It's free. It's small. It's the first time we're showing the cluster.'",                                  "Draft", "Sara"),
    (date(2026,5,13), "Wed", "Wk 2 (11–17 May)", "Expo signup (peer)",       "Slack",    "Expo signup is live — here's who it's for",                                          "Cross-post the signup link inside the L&D Shakers community. Lead with the question the expo answers: 'Is this safe for someone like me?' Tag the practitioners who'll be there.",                                "Draft", "Sara"),
    (date(2026,5,15), "Fri", "Wk 2 (11–17 May)", "Concept teaser",           "LinkedIn", "Meet the Override Log — the one practice every cluster member keeps",                "Carousel or single-image post explaining the three-field practice (AI suggested / I changed / Because). Position as governance-friendly AND craft-friendly.",                                                       "Draft", "Sara"),

    (date(2026,5,18), "Mon", "Wk 3 (18–24 May)", "Practitioner spotlight #1","LinkedIn", "Meet [Practitioner 1] — what they protected when AI entered their practice",         "Quote-led post: a sentence in the practitioner's voice + a one-line frame from us. Link to expo signup at the end.",                                                                                              "Draft", "Sara"),
    (date(2026,5,20), "Wed", "Wk 3 (18–24 May)", "Practitioner spotlight #2","Slack",    "[Practitioner 2] is bringing their case to the expo — here's the question they're holding", "Inside the community: share the messy version of the question, not the polished one. Invite reactions in-thread.",                                                                                              "Draft", "Sara"),
    (date(2026,5,22), "Fri", "Wk 3 (18–24 May)", "Cluster vs. course",       "LinkedIn", "We chose 'cluster', not 'course'. Here's the difference — and why it matters for L&D", "Short essay-style post. Use the Sprint/Studio choice and the bi-weekly cycle as proof points. End with expo signup CTA.",                                                                                         "Draft", "Sara"),

    (date(2026,5,25), "Mon", "Wk 4 (25–31 May)", "Practitioner cases drop",  "LinkedIn", "4 practitioners. 4 real cases. Read them before the expo.",                                "Headline drop with link to the published cases. Hero image of all four practitioners. Frame: 'These aren't success stories. They're working stories — with the override decisions visible.'",                  "Draft", "Sara"),
    (date(2026,5,25), "Mon", "Wk 4 (25–31 May)", "Practitioner cases drop",  "Slack",    "Cases are live — drop your reactions in the thread before Tuesday",                  "In-community drop with all four cases. Pin the message. Ask a single discussion question: 'Which override decision do you most want to interrogate at the expo?'",                                                "Draft", "Sara"),
    (date(2026,5,27), "Wed", "Wk 4 (25–31 May)", "Practitioner spotlight #3","LinkedIn", "[Practitioner 3]: from personal pilot to team practice — what changed",              "Lift the most quotable line from their published case. Link case + expo signup.",                                                                                                                                  "Draft", "Sara"),
    (date(2026,5,28), "Thu", "Wk 4 (25–31 May)", "Practitioner spotlight #4 + countdown", "Slack", "[Practitioner 4]'s case in 90 seconds + 5 days to expo",                       "Short summary + link to full case + signup nudge for community members who haven't registered.",                                                                                                                   "Draft", "Sara"),
    (date(2026,5,29), "Fri", "Wk 4 (25–31 May)", "Expo final push",          "LinkedIn", "Last call for Tuesday's expo — here's what we'll actually do in the room",          "Walk through the 60 minutes: what attendees will hear, see, and leave with. Address the unspoken question: 'Is this for me?'",                                                                                    "Draft", "Sara"),

    (date(2026,6,1),  "Mon", "Wk 5 (1–7 June)",  "Expo eve",                 "LinkedIn", "Tomorrow at 17:00 CEST — what to bring (and what to leave behind)",                  "Warm tone. Short. 'Bring: one decision you're stuck on. Leave behind: the pressure to have an AI strategy.' Final signup link.",                                                                                   "Draft", "Sara"),
    (date(2026,6,2),  "Tue", "Wk 5 (1–7 June)",  "Day-of expo reminder",     "Slack",    "Today 17:00 CEST — Zoom link inside",                                                "Pin in #ai-cluster. Zoom link, agenda, and the one Override Log we'll open the room with.",                                                                                                                       "Draft", "Sara"),
    (date(2026,6,2),  "Tue", "Wk 5 (1–7 June)",  "Cohort open announcement", "LinkedIn", "The cohort is open. 12 seats. Starts Tue 9 June, 17:00 CEST.",                            "Posted POST-expo, same evening. Lead with the Expo's signal moment. Cap, dates, link, and one sentence on who it's for.",                                                                                          "Draft", "Sara"),
    (date(2026,6,3),  "Wed", "Wk 5 (1–7 June)",  "Post-expo recap + cohort signup", "Slack", "Couldn't make it? 6-min recap + cohort signup link",                                  "Recap thread for community members who missed it: 3 quotes, 1 practitioner clip, 1 sentence on next step. Direct link to cohort signup.",                                                                          "Draft", "Sara"),
    (date(2026,6,5),  "Fri", "Wk 5 (1–7 June)",  "Cohort signup push",       "LinkedIn", "'Is this safe for someone like me?' — what we built into the cluster",              "Objection-handling post. Address the two prior questions head-on (compliance/governance + craft/voice). End with cohort signup link — 'Closes Sunday.'",                                                      "Draft", "Sara"),
    (date(2026,6,7),  "Sun", "Wk 5 (1–7 June)",  "Final cohort signup push", "Slack",    "Last seats for Tuesday's kickoff — final call by Sunday EOD",                        "Short. Honest. Last-call energy. Tag the practitioners who are joining as anchor members.",                                                                                                                       "Draft", "Sara"),

    (date(2026,6,8),  "Mon", "Wk 6 (8–9 June)",  "Cohort kickoff prep",      "Slack",    "Tomorrow 17:00 CEST — what to bring to session 1",                                  "Pinned in the cohort channel. Pre-read link, calendar invite check, and the opening prompt: 'What do you most need to protect?'",                                                                                  "Draft", "Sara"),
    (date(2026,6,9),  "Tue", "Wk 6 (8–9 June)",  "Cohort live",              "LinkedIn", "And… we're live. Cohort 1 of the AI Learning Cluster starts tonight.",              "Celebratory post. Name the cohort members (with permission). Thank the community. Signal: 'Cohort 2 will open in autumn — join the waitlist.'",                                                              "Draft", "Sara"),
]

wb = Workbook()
ws = wb.active
ws.title = "Campaign Calendar"

headers = ["Date", "Day", "Week", "Content type", "Platform", "Topic / Angle", "Hook / Copy hint", "Status", "Owner"]
ws.append(headers)

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="2C3E50")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="D5D8DC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = border

body_font = Font(name="Arial", size=10)
body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

milestone_dates = {date(2026,5,25), date(2026,6,2), date(2026,6,9)}

for r in rows:
    ws.append(r)
    row_idx = ws.max_row
    is_milestone = r[0] in milestone_dates
    for col_idx in range(1, len(headers)+1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = Font(name="Arial", size=10, bold=is_milestone)
        c.alignment = body_align
        c.border = border
    ws.cell(row=row_idx, column=1).number_format = "ddd d-mmm"
    if r[4] == "LinkedIn":
        ws.cell(row=row_idx, column=5).fill = PatternFill("solid", start_color="DEEBF7")
    elif r[4] == "Slack":
        ws.cell(row=row_idx, column=5).fill = PatternFill("solid", start_color="F4ECF7")
    if is_milestone:
        for col_idx in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", start_color="FFF3CD")
        ws.cell(row=row_idx, column=5).fill = PatternFill("solid", start_color="FFE08A")

widths = {"A":12, "B":6, "C":18, "D":24, "E":11, "F":42, "G":60, "H":12, "I":10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

for row_idx in range(2, ws.max_row+1):
    ws.row_dimensions[row_idx].height = 60

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

status_col_letter = "H"
status_range = f"{status_col_letter}2:{status_col_letter}{ws.max_row}"
draft_fill     = PatternFill("solid", start_color="F9E79F")
scheduled_fill = PatternFill("solid", start_color="AED6F1")
live_fill      = PatternFill("solid", start_color="A9DFBF")
ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Draft"'],     fill=draft_fill))
ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Scheduled"'], fill=scheduled_fill))
ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"Live"'],      fill=live_fill))

# Legend / notes sheet
ws2 = wb.create_sheet("Legend & Notes")
notes = [
    ("L&D Shakers — AI Learning Cluster Launch Campaign", ""),
    ("Window", "5 May 2026 → 9 June 2026"),
    ("Cadence", "Light: 1 LinkedIn + 1 Slack/week, plus milestone posts"),
    ("Tone", "Warm, community-led, peer-to-peer"),
    ("", ""),
    ("Key milestones", ""),
    ("Mon 25 May", "Practitioner cases published — LinkedIn drop + Slack drop"),
    ("Tue 2 Jun, 17:00 CEST", "Opening Expo + cohort open announcement (same evening)"),
    ("Tue 9 Jun, 17:00 CEST", "Cohort kickoff (note: 9 June 2026 is a Tuesday, not Monday)"),
    ("", ""),
    ("Status colours", ""),
    ("Draft", "Yellow"),
    ("Scheduled", "Blue"),
    ("Live", "Green"),
    ("", ""),
    ("Platform colours", ""),
    ("LinkedIn", "Light blue cell"),
    ("Slack", "Light purple cell"),
    ("Milestone row", "Amber highlight + bold text"),
    ("", ""),
    ("Practitioner placeholders", "Replace [Practitioner 1–4] with confirmed names once locked. Spotlights are on 18 May, 20 May, 27 May, 28 May."),
]
for row in notes:
    ws2.append(row)
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 80
for r_i in range(1, ws2.max_row+1):
    for c_i in (1,2):
        c = ws2.cell(row=r_i, column=c_i)
        c.font = Font(name="Arial", size=10, bold=(c_i==1))
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws2["A1"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", start_color="2C3E50")
ws2["B1"].fill = PatternFill("solid", start_color="2C3E50")

out = "/sessions/festive-awesome-dirac/mnt/ai-transform-space/ld-shakers-ai-cluster-launch-calendar.xlsx"
wb.save(out)
print("Saved:", out, "rows:", ws.max_row-1)
